"""
报告内容提取器
支持从PDF和HTML文件中提取文本内容
"""

import PyPDF2
import pdfplumber
from bs4 import BeautifulSoup
import logging
import os

logger = logging.getLogger(__name__)


class ReportContentExtractor:
    """提取PDF和HTML报告的内容"""

    def extract_from_pdf(self, pdf_path):
        """从PDF提取全文文本

        Args:
            pdf_path: PDF文件路径

        Returns:
            str: 提取的文本内容
        """
        if not os.path.exists(pdf_path):
            logger.error(f"PDF file not found: {pdf_path}")
            return ""

        try:
            # 优先使用pdfplumber（更准确的文本提取）
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text.strip()
        except Exception as e:
            logger.warning(f"pdfplumber extraction failed for {pdf_path}: {e}")
            # 备用方案：PyPDF2
            try:
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text += page_text + "\n"
                    return text.strip()
            except Exception as e2:
                logger.error(f"PyPDF2 extraction also failed for {pdf_path}: {e2}")
                return ""

    def extract_from_html(self, html_path):
        """从HTML文件提取内容

        Args:
            html_path: HTML文件路径

        Returns:
            str: 提取的文本内容
        """
        if not os.path.exists(html_path):
            logger.error(f"HTML file not found: {html_path}")
            return ""

        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                html_content = f.read()

            soup = BeautifulSoup(html_content, 'html.parser')

            # 移除脚本和样式标签
            for script in soup(["script", "style", "nav", "footer", "header"]):
                script.decompose()

            # 获取文本内容
            text = soup.get_text(separator='\n', strip=True)

            # 清理多余空行
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            return '\n'.join(lines)

        except Exception as e:
            logger.error(f"HTML extraction failed for {html_path}: {e}")
            return ""

    def extract_from_excel(self, excel_path):
        """从Excel文件提取内容

        Args:
            excel_path: Excel文件路径 (.xlsx, .xls)

        Returns:
            str: 提取的文本内容
        """
        if not os.path.exists(excel_path):
            logger.error(f"Excel file not found: {excel_path}")
            return ""

        try:
            # 优先使用 openpyxl
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            text_content = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")

                # 提取所有非空行
                for row in sheet.iter_rows(values_only=True):
                    # 过滤 None 值和空字符串
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    row_text = '  '.join([v for v in row_values if v.strip()])
                    if row_text:
                        text_content.append(row_text)

            wb.close()
            return '\n'.join(text_content)

        except ImportError:
            logger.warning("openpyxl not installed, trying pandas fallback")
        except Exception as e:
            logger.warning(f"openpyxl extraction failed for {excel_path}: {e}")

        # 后备方案：pandas
        try:
            import pandas as pd
            all_sheets = []
            for sheet_name in pd.ExcelFile(excel_path).sheet_names:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                sheet_text = df.to_string(index=False, na_rep='')
                all_sheets.append(f"=== Sheet: {sheet_name} ===\n{sheet_text}")
            return '\n\n'.join(all_sheets)
        except Exception as e:
            logger.error(f"Pandas Excel extraction also failed for {excel_path}: {e}")
            return ""

    def extract_content(self, file_path):
        """根据文件类型自动提取内容

        Args:
            file_path: 文件路径

        Returns:
            str: 提取的文本内容
        """
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            return ""

        file_lower = file_path.lower()

        if file_lower.endswith('.pdf'):
            return self.extract_from_pdf(file_path)
        elif file_lower.endswith(('.xlsx', '.xls')):
            return self.extract_from_excel(file_path)
        elif file_lower.endswith(('.html', '.htm')):
            return self.extract_from_html(file_path)
        elif file_lower.endswith('.txt'):
            # 纯文本文件直接读取
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except Exception as e:
                logger.error(f"Text file reading failed for {file_path}: {e}")
                return ""
        else:
            logger.warning(f"Unsupported file type: {file_path}")
            return ""

    def get_file_info(self, file_path):
        """获取文件信息

        Args:
            file_path: 文件路径

        Returns:
            dict: 包含文件大小、类型等信息
        """
        if not os.path.exists(file_path):
            return {'exists': False}

        stat_info = os.stat(file_path)
        file_ext = os.path.splitext(file_path)[1].lower()

        return {
            'exists': True,
            'size_bytes': stat_info.st_size,
            'size_mb': round(stat_info.st_size / (1024 * 1024), 2),
            'extension': file_ext,
            'type': self._get_file_type_name(file_ext)
        }

    def _get_file_type_name(self, extension):
        """根据扩展名返回文件类型名称"""
        type_map = {
            '.pdf': 'PDF Document',
            '.xlsx': 'Excel Spreadsheet',
            '.xls': 'Excel Spreadsheet',
            '.html': 'HTML Document',
            '.htm': 'HTML Document',
            '.txt': 'Text File'
        }
        return type_map.get(extension, 'Unknown')
